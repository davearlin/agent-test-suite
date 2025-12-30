from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, status, BackgroundTasks
from fastapi.responses import Response
from sqlalchemy.orm import Session, joinedload
import logging

from app.core.database import get_db
from app.core.validation import validate_session_parameters
from app.core.csv_utils import escape_csv_value
from app.api.auth import get_current_user
from app.models import User, TestRun, Dataset, Question, TestResult, TestRunDataset, EvaluationParameter, TestRunEvaluationConfig, TestResultParameterScore
from app.models.schemas import (
    TestRun as TestRunSchema,
    TestRunRead,
    TestRunCreate,
    TestRunUpdate,
    TestResult as TestResultSchema,
    TestProgress
)
from app.services.test_execution_service import TestRunExecutionService

logger = logging.getLogger(__name__)

router = APIRouter()


@router.get("/", response_model=List[TestRunRead])
async def list_test_runs(
    skip: int = 0,
    limit: int = 100,
    dataset_id: Optional[int] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """List test runs with optional filtering.
    
    NOTE: This endpoint intentionally does NOT load test_results or full dataset questions
    to keep payloads small. Returns DatasetSummary instead of full Dataset.
    """
    from sqlalchemy.orm import noload, selectinload
    from sqlalchemy import func
    from app.models.schemas import DatasetSummary
    
    # Join with User table to get creator information
    # Load datasets but not their questions
    query = db.query(TestRun).options(
        joinedload(TestRun.created_by),
        noload(TestRun.test_results),  # Don't load test results in list view
        selectinload(TestRun.datasets).noload(Dataset.questions),  # Load datasets but not questions
        selectinload(TestRun.datasets).joinedload(Dataset.owner),  # Load dataset owner for owner_name
        noload(TestRun.dataset)  # Don't load legacy single dataset
    )
    
    if dataset_id:
        query = query.filter(TestRun.dataset_id == dataset_id)
    
    if status:
        query = query.filter(TestRun.status == status)
    
    # Non-admin users can only see their own test runs
    if current_user.role != "admin":
        query = query.filter(TestRun.created_by_id == current_user.id)
    
    test_runs = query.order_by(TestRun.created_at.desc()).offset(skip).limit(limit).all()
    
    # Convert to schemas and manually populate user fields and dataset summaries
    result = []
    for test_run in test_runs:
        try:
            # Convert datasets to lightweight summaries FIRST (exclude questions to reduce response size)
            dataset_summaries = []
            if test_run.datasets:
                for ds in test_run.datasets:
                    # Get question count
                    question_count = db.query(func.count(Question.id)).filter(
                        Question.dataset_id == ds.id
                    ).scalar() or 0
                    
                    dataset_summaries.append({
                        'id': ds.id,
                        'name': ds.name,
                        'category': ds.category,
                        'version': ds.version,
                        'question_count': question_count,
                        'created_at': ds.created_at,
                        'owner_name': ds.owner.full_name if hasattr(ds, 'owner') and ds.owner else 'Unknown'
                    })
            
            # Create a copy of test_run with converted datasets
            # Use getattr with defaults to handle old test runs that may not have all fields
            test_run_dict = {
                'id': test_run.id,
                'name': getattr(test_run, 'name', 'Unnamed Test'),
                'description': getattr(test_run, 'description', None),
                'project_id': getattr(test_run, 'project_id', None),
                'agent_id': getattr(test_run, 'agent_id', None),
                'agent_name': getattr(test_run, 'agent_name', ''),
                'agent_display_name': getattr(test_run, 'agent_display_name', None),
                'flow_name': getattr(test_run, 'flow_name', 'Default Start Flow'),
                'flow_display_name': getattr(test_run, 'flow_display_name', None),
                'page_name': getattr(test_run, 'page_name', 'Start Page'),
                'page_display_name': getattr(test_run, 'page_display_name', None),
                'environment': getattr(test_run, 'environment', 'draft'),
                'playbook_id': getattr(test_run, 'playbook_id', None),
                'playbook_display_name': getattr(test_run, 'playbook_display_name', None),
                'session_parameters': getattr(test_run, 'session_parameters', None),
                'pre_prompt_messages': getattr(test_run, 'pre_prompt_messages', None),
                'post_prompt_messages': getattr(test_run, 'post_prompt_messages', None),
                'enable_webhook': getattr(test_run, 'enable_webhook', True),
                'evaluation_parameters': getattr(test_run, 'evaluation_parameters', None),
                'llm_model_id': getattr(test_run, 'llm_model_id', None),
                'evaluation_model_id': getattr(test_run, 'evaluation_model_id', ''),
                'batch_size': getattr(test_run, 'batch_size', 10),
                'created_at': test_run.created_at,
                'created_by_id': test_run.created_by_id,
                'status': test_run.status,
                'total_questions': getattr(test_run, 'total_questions', 0),
                'completed_questions': getattr(test_run, 'completed_questions', 0),
                'average_score': getattr(test_run, 'average_score', None),
                'started_at': getattr(test_run, 'started_at', None),
                'completed_at': getattr(test_run, 'completed_at', None),
                'datasets': dataset_summaries
            }
            
            # Add user information
            if test_run.created_by:
                test_run_dict['created_by_email'] = test_run.created_by.email
                test_run_dict['created_by_name'] = test_run.created_by.full_name
            else:
                test_run_dict['created_by_email'] = 'unknown@example.com'
                test_run_dict['created_by_name'] = 'Unknown User'
            
            result.append(TestRunRead(**test_run_dict))
        except Exception as e:
            # Log validation errors but don't fail the entire request
            import logging
            logging.error(f"Warning: Could not validate test run {test_run.id}: {str(e)}")
            # Skip this test run rather than failing the entire request
            continue
    
    return result


@router.post("/", response_model=TestRunSchema)
async def create_test_run(
    test_run_data: TestRunCreate,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Create and start a new test run (supports both single and multiple datasets)."""
    
    if current_user.role not in ["admin", "test_manager"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions"
        )
    
    # Determine which datasets to use
    dataset_ids = []
    if test_run_data.dataset_ids:
        # Multi-dataset test run
        dataset_ids = test_run_data.dataset_ids
    elif test_run_data.dataset_id:
        # Single dataset test run (backward compatibility)
        dataset_ids = [test_run_data.dataset_id]
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Either dataset_id or dataset_ids must be provided"
        )
    
    # Verify all datasets exist and user has access
    datasets = []
    total_questions = 0
    
    for dataset_id in dataset_ids:
        dataset = db.query(Dataset).filter(Dataset.id == dataset_id).first()
        if not dataset:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Dataset with ID {dataset_id} not found"
            )
        
        if current_user.role != "admin" and dataset.owner_id != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=f"Not enough permissions to use dataset '{dataset.name}'"
            )
        
        # Count questions in this dataset
        question_count = db.query(Question).filter(Question.dataset_id == dataset.id).count()
        if question_count == 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Dataset '{dataset.name}' has no questions to test"
            )
        
        datasets.append(dataset)
        total_questions += question_count
    
    # Validate session parameters for duplicates and format
    validated_session_parameters = validate_session_parameters(
        test_run_data.session_parameters, 
        context="test run"
    )

    # Debug log for pre/post prompt configuration
    print(f"🔧 Creating test run with:")
    print(f"   - Pre-prompt messages: {test_run_data.pre_prompt_messages}")
    print(f"   - Post-prompt messages: {test_run_data.post_prompt_messages}")  
    print(f"   - Enable webhook: {test_run_data.enable_webhook}")

    # Create test run
    db_test_run = TestRun(
        name=test_run_data.name,
        description=test_run_data.description,
        dataset_id=test_run_data.dataset_id,  # For backward compatibility
        created_by_id=current_user.id,
        project_id=test_run_data.project_id,
        agent_id=test_run_data.agent_id,
        agent_name=test_run_data.agent_name,  # Full path for API calls
        agent_display_name=test_run_data.agent_display_name,  # Display name for UI
        flow_name=test_run_data.flow_name,
        flow_display_name=test_run_data.flow_display_name,  # Display name for UI
        page_name=test_run_data.page_name,
        page_display_name=test_run_data.page_display_name,  # Display name for UI
        environment=test_run_data.environment,
        batch_size=test_run_data.batch_size,
        playbook_id=test_run_data.playbook_id,  # Playbook support
        playbook_display_name=test_run_data.playbook_display_name,  # Playbook display name
        llm_model_id=test_run_data.llm_model_id,  # LLM model for playbooks
        evaluation_model_id=test_run_data.evaluation_model_id,  # LLM model for evaluation
        session_parameters=validated_session_parameters,  # Session parameters (validated)
        enable_webhook=test_run_data.enable_webhook,  # Webhook configuration
        pre_prompt_messages=test_run_data.pre_prompt_messages,  # Pre-prompt messages
        post_prompt_messages=test_run_data.post_prompt_messages,  # Post-prompt messages
        total_questions=total_questions,
        status="pending"
    )
    
    db.add(db_test_run)
    db.commit()
    db.refresh(db_test_run)
    
    # Create junction table entries for multi-dataset support
    from app.models import TestRunDataset
    for dataset in datasets:
        test_run_dataset = TestRunDataset(
            test_run_id=db_test_run.id,
            dataset_id=dataset.id
        )
        db.add(test_run_dataset)
    
    # Create evaluation parameter configuration if provided, or use defaults
    if test_run_data.evaluation_parameters:
        # Verify all evaluation parameters exist
        param_ids = [param_config.parameter_id for param_config in test_run_data.evaluation_parameters]
        existing_params = db.query(EvaluationParameter).filter(
            EvaluationParameter.id.in_(param_ids)
        ).all()
        
        if len(existing_params) != len(param_ids):
            existing_ids = [p.id for p in existing_params]
            missing_ids = [pid for pid in param_ids if pid not in existing_ids]
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Evaluation parameters not found: {missing_ids}"
            )
        
        # Convert parameter configs to JSON format for storage
        parameters_json = [
            {
                "parameter_id": param_config.parameter_id,
                "weight": param_config.weight,
                "enabled": param_config.enabled
            }
            for param_config in test_run_data.evaluation_parameters
        ]
        
        evaluation_config = TestRunEvaluationConfig(
            test_run_id=db_test_run.id,
            user_id=current_user.id,
            name=f"Config for {test_run_data.name}",
            is_default=False,
            parameters=parameters_json
        )
        db.add(evaluation_config)
    else:
        # Create default evaluation configuration using all active evaluation parameters
        active_params = db.query(EvaluationParameter).filter(
            EvaluationParameter.is_active == True
        ).all()
        
        if active_params:
            # Distribute weights evenly among active parameters
            weight_per_param = 100 // len(active_params)
            remaining_weight = 100 % len(active_params)
            
            parameters_json = []
            for i, param in enumerate(active_params):
                # Give the first parameter any remaining weight due to rounding
                weight = weight_per_param + (remaining_weight if i == 0 else 0)
                parameters_json.append({
                    "parameter_id": param.id,
                    "weight": weight,
                    "enabled": True
                })
            
            evaluation_config = TestRunEvaluationConfig(
                test_run_id=db_test_run.id,
                user_id=current_user.id,
                name=f"Default Config for {test_run_data.name}",
                is_default=True,
                parameters=parameters_json
            )
            db.add(evaluation_config)
    
    db.commit()
    
    # Start test execution in background
    service = TestRunExecutionService(user=current_user, db=db)
    background_tasks.add_task(service.execute_test_run, db_test_run.id)
    
    return db_test_run


@router.get("/{test_run_id}", response_model=TestRunSchema)
async def get_test_run(
    test_run_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get a specific test run by ID.
    
    NOTE: This endpoint does NOT load test_results to keep payload manageable.
    Use GET /tests/{test_run_id}/results with pagination to fetch test results.
    """
    from sqlalchemy.orm import noload
    
    test_run = db.query(TestRun)\
        .options(
            joinedload(TestRun.evaluation_config),
            noload(TestRun.test_results)  # Don't load all test results - use /results endpoint instead
        )\
        .filter(TestRun.id == test_run_id)\
        .first()
    
    if not test_run:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Test run not found"
        )
    
    # Check permissions
    if current_user.role != "admin" and test_run.created_by_id != current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions"
        )
    
    # If we have a playbook_id but missing or incorrect playbook_display_name, fetch it
    if test_run.playbook_id:
        needs_update = (
            not test_run.playbook_display_name or 
            test_run.playbook_display_name == test_run.playbook_id or
            # Check if it's still a GUID (means it wasn't properly converted)
            (len(test_run.playbook_display_name) > 30 and 
             any(c in test_run.playbook_display_name for c in ['-', 'a', 'b', 'c', 'd', 'e', 'f']) and
             not any(c.isupper() for c in test_run.playbook_display_name[:10]))
        )
        
        if needs_update:
            try:
                from app.services.dialogflow_service import DialogflowService
                df_service = DialogflowService(user=current_user, db=db)
                playbooks = await df_service.list_playbooks()
                
                logger.info(f"Fetching playbook display name for: {test_run.playbook_id}")
                logger.info(f"Current display name: {test_run.playbook_display_name}")
                logger.info(f"Available playbooks: {[p.get('name', '') for p in playbooks]}")
                
                for playbook in playbooks:
                    playbook_name = playbook.get('name', '')
                    # Check both full path ending and just the GUID part
                    if (playbook_name.endswith(test_run.playbook_id) or 
                        test_run.playbook_id in playbook_name):
                        old_name = test_run.playbook_display_name
                        test_run.playbook_display_name = playbook.get('displayName', test_run.playbook_id)
                        logger.info(f"Updated playbook display name from '{old_name}' to '{test_run.playbook_display_name}'")
                        # Update the database record
                        db.commit()
                        break
                else:
                    logger.warning(f"No matching playbook found for ID: {test_run.playbook_id}")
            except Exception as e:
                # If we can't fetch the playbook name, keep the existing value
                logger.warning(f"Could not fetch playbook display name: {e}")
    
    return test_run


@router.put("/{test_run_id}", response_model=TestRunSchema)
async def update_test_run(
    test_run_id: int,
    test_run_update: TestRunUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Update a test run (mainly for status changes)."""
    test_run = db.query(TestRun).filter(TestRun.id == test_run_id).first()
    
    if not test_run:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Test run not found"
        )
    
    # Check permissions
    if current_user.role not in ["admin"] and test_run.created_by_id != current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions"
        )
    
    # Update fields
    update_data = test_run_update.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(test_run, field, value)
    
    db.commit()
    db.refresh(test_run)
    
    return test_run


@router.delete("/{test_run_id}")
async def delete_test_run(
    test_run_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete a test run and all its results."""
    test_run = db.query(TestRun).filter(TestRun.id == test_run_id).first()
    
    if not test_run:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Test run not found"
        )
    
    # Check permissions
    if current_user.role not in ["admin"] and test_run.created_by_id != current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions"
        )
    
    db.delete(test_run)
    db.commit()
    
    return {"message": "Test run deleted successfully"}


@router.get("/{test_run_id}/results", response_model=List[TestResultSchema])
async def get_test_results(
    test_run_id: int,
    skip: int = 0,
    limit: int = 1000,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get all results for a specific test run with pagination.
    
    Default limit is 1000 to handle most test runs in a single request while
    still preventing issues with extremely large test runs (5000+ questions).
    """
    test_run = db.query(TestRun).filter(TestRun.id == test_run_id).first()
    
    if not test_run:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Test run not found"
        )
    
    # Check permissions
    if current_user.role != "admin" and test_run.created_by_id != current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions"
        )
    
    results = (
        db.query(TestResult)
        .options(
            joinedload(TestResult.parameter_scores).joinedload(TestResultParameterScore.parameter),
            joinedload(TestResult.question)
        )
        .filter(TestResult.test_run_id == test_run_id)
        .offset(skip)
        .limit(limit)
        .all()
    )
    
    return results


@router.post("/{test_run_id}/cancel")
async def cancel_test_run(
    test_run_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Cancel a running test run."""
    test_run = db.query(TestRun).filter(TestRun.id == test_run_id).first()
    
    if not test_run:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Test run not found"
        )
    
    # Check permissions
    if current_user.role not in ["admin"] and test_run.created_by_id != current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions"
        )
    
    if test_run.status not in ["pending", "running"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Test run cannot be cancelled in its current state"
        )
    
    test_run.status = "cancelled"
    db.commit()
    
    return {"message": "Test run cancelled successfully"}


@router.post("/integration-test")
async def test_integration(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Test endpoint to verify Dialogflow and LLM services are working."""
    
    if current_user.role not in ["admin"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin access required for integration testing"
        )
    
    try:
        from app.services.dialogflow_service import DialogflowService
        from app.services.llm_judge_service import LLMJudgeService
        
        # Test Dialogflow service
        df_service = DialogflowService(user=current_user, db=db)
        
        # Test agent listing
        agents = await df_service.list_agents()
        df_status = {
            "service": "Dialogflow",
            "available": df_service.is_available(),
            "agents_found": len(agents),
            "agents": agents[:3] if agents else []  # Return first 3 for testing
        }
        
        # Test intent detection
        test_response = await df_service.detect_intent(
            agent_name="projects/test/locations/us-central1/agents/test-agent",
            session_id="integration-test-session",
            text_input="Hello, can you help me?"
        )
        
        df_status["intent_detection"] = {
            "success": True,
            "response_time_ms": test_response.get("response_time_ms", 0),
            "is_mock": test_response.get("is_mock", False),
            "confidence": test_response.get("intent_detection_confidence", 0)
        }
        
        # Test LLM Judge service
        llm_judge_service = LLMJudgeService(user=current_user, db=db)
        
        llm_test = await llm_judge_service.evaluate_response(
            expected_answer="I can help you with payroll questions.",
            actual_answer="Hello! I'm here to assist with your payroll needs.",
            question="Can you help me with payroll?",
            detect_empathy=False,
            no_match_expected=False
        )
        
        llm_status = {
            "service": "LLM Judge", 
            "available": True,
            "evaluation_test": {
                "similarity_score": llm_test.get("similarity_score", 0),
                "has_reasoning": bool(llm_test.get("evaluation_reasoning")),
                "error": llm_test.get("error")
            }
        }
        
        return {
            "status": "success",
            "timestamp": "2025-09-10T00:00:00Z",
            "services": {
                "dialogflow": df_status,
                "llm_judge": llm_status
            },
            "overall_health": "Both core services are functional"
        }
        
    except Exception as e:
        return {
            "status": "error",
            "timestamp": "2025-09-10T00:00:00Z", 
            "error": str(e),
            "overall_health": "Integration test failed"
        }


@router.get("/{test_run_id}/export")
async def export_test_run(
    test_run_id: int,
    format: str = "csv",
    timezone: str = "UTC",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export test run results to CSV or Excel format.
    
    Args:
        format: Export format - 'csv' or 'excel' (default: csv)
        timezone: IANA timezone for date formatting in Excel (e.g., 'America/New_York', 'Europe/London')
    """
    # Get the test run
    test_run = db.query(TestRun).filter(TestRun.id == test_run_id).first()
    
    if not test_run:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Test run not found"
        )
    
    # Check permission
    if current_user.role != "admin" and test_run.created_by_id != current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized to access this test run"
        )
    
    # Get test results with all related data
    test_results = db.query(TestResult).options(
        joinedload(TestResult.question),
        joinedload(TestResult.parameter_scores).joinedload(TestResultParameterScore.parameter)
    ).filter(TestResult.test_run_id == test_run_id).all()
    
    if not test_results:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No test results found for this test run"
        )
    
    # Generate filename base
    safe_name = test_run.name.replace(' ', '_').lower()
    safe_name = ''.join(c for c in safe_name if c.isalnum() or c in ('_', '-'))
    
    # Prepare data rows
    rows_data = []
    for result in test_results:
        # Get parameter scores sorted by parameter_id
        parameter_scores = sorted(result.parameter_scores, key=lambda x: x.parameter_id) if result.parameter_scores else []
        
        # Calculate overall score as weighted average from parameter scores (same as frontend)
        if parameter_scores:
            total_weight = sum(ps.weight_used or 0 for ps in parameter_scores)
            if total_weight > 0:
                weighted_sum = sum((ps.score or 0) * (ps.weight_used or 0) for ps in parameter_scores)
                overall_score = round(weighted_sum / total_weight, 1)
            else:
                overall_score = 0
        else:
            # Fallback to stored overall_score or similarity_score if no parameter scores
            overall_score = result.overall_score if result.overall_score is not None else (
                result.similarity_score if result.similarity_score is not None else 0
            )
        
        # Get actual answer from various response fields
        actual_answer = (
            result.actual_answer or 
            (result.dialogflow_response.get('response_text') if result.dialogflow_response else None) or
            (result.dialogflow_response.get('fulfillment_text') if result.dialogflow_response else None) or
            (f"Error: {result.error_message}" if result.error_message else "No response")
        )
        
        row = {
            'question': result.question.question_text if result.question else '',
            'expected_answer': result.question.expected_answer if result.question else '',
            'actual_answer': actual_answer,
            'overall_score': overall_score,
            'parameter_scores': parameter_scores
        }
        rows_data.append(row)
    
    if format.lower() == 'excel':
        return _create_excel_export(test_run, rows_data, safe_name, timezone)
    else:
        return _create_csv_export(test_run, rows_data, safe_name)


def _create_csv_export(test_run, rows_data, safe_name):
    """Create CSV export response."""
    headers = [
        'Question', 'Expected Answer', 'Actual Answer', 
        'Overall Score (%)',
        'Parameter 1 Name', 'Parameter 1 Score (%)', 'Parameter 1 Weight (%)', 'Parameter 1 Reasoning',
        'Parameter 2 Name', 'Parameter 2 Score (%)', 'Parameter 2 Weight (%)', 'Parameter 2 Reasoning',
        'Parameter 3 Name', 'Parameter 3 Score (%)', 'Parameter 3 Weight (%)', 'Parameter 3 Reasoning',
        'Parameter 4 Name', 'Parameter 4 Score (%)', 'Parameter 4 Weight (%)', 'Parameter 4 Reasoning',
        'Parameter 5 Name', 'Parameter 5 Score (%)', 'Parameter 5 Weight (%)', 'Parameter 5 Reasoning'
    ]
    
    csv_rows = [','.join(escape_csv_value(header) for header in headers)]
    
    for row in rows_data:
        base_fields = [
            escape_csv_value(row['question']),
            escape_csv_value(row['expected_answer']),
            escape_csv_value(row['actual_answer']),
            str(row['overall_score'])
        ]
        
        parameter_fields = []
        parameter_scores = row['parameter_scores']
        
        for i in range(5):
            if i < len(parameter_scores):
                param_score = parameter_scores[i]
                parameter_fields.extend([
                    escape_csv_value(param_score.parameter.name if param_score.parameter else f'Parameter {param_score.parameter_id}'),
                    str(param_score.score or 0),
                    str(param_score.weight_used or 0),
                    escape_csv_value(param_score.reasoning or '')
                ])
            else:
                parameter_fields.extend(['', '', '', ''])
        
        row_data = base_fields + parameter_fields
        csv_rows.append(','.join(row_data))
    
    csv_content = '\n'.join(csv_rows)
    filename = f"test-run-{test_run.id}-{safe_name}-results.csv"
    
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def _create_excel_export(test_run, rows_data, safe_name, timezone: str = "UTC"):
    """Create professionally styled Excel export response."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from zoneinfo import ZoneInfo
    from datetime import datetime
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Alternating row colors
    row_fill_light = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    row_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Border style
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    
    # Score color fills (red -> yellow -> green gradient)
    def get_score_fill(score):
        """Get fill color based on score (0-100)."""
        if score >= 80:
            return PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
        elif score >= 60:
            return PatternFill(start_color="D4EDBC", end_color="D4EDBC", fill_type="solid")  # Light green
        elif score >= 40:
            return PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
        elif score >= 20:
            return PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
        else:
            return PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Red
    
    def get_score_font(score):
        """Get font color based on score."""
        if score >= 60:
            return Font(bold=True, color="006100")  # Dark green
        elif score >= 40:
            return Font(bold=True, color="9C5700")  # Dark yellow/brown
        else:
            return Font(bold=True, color="9C0006")  # Dark red
    
    # Headers
    headers = [
        'Question', 'Expected Answer', 'Actual Answer', 
        'Overall Score (%)',
        'Parameter 1 Name', 'Parameter 1 Score (%)', 'Parameter 1 Weight (%)', 'Parameter 1 Reasoning',
        'Parameter 2 Name', 'Parameter 2 Score (%)', 'Parameter 2 Weight (%)', 'Parameter 2 Reasoning',
        'Parameter 3 Name', 'Parameter 3 Score (%)', 'Parameter 3 Weight (%)', 'Parameter 3 Reasoning',
        'Parameter 4 Name', 'Parameter 4 Score (%)', 'Parameter 4 Weight (%)', 'Parameter 4 Reasoning',
        'Parameter 5 Name', 'Parameter 5 Score (%)', 'Parameter 5 Weight (%)', 'Parameter 5 Reasoning'
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data rows
    for row_idx, row in enumerate(rows_data, 2):
        # Determine row fill (alternating)
        row_fill = row_fill_light if row_idx % 2 == 0 else row_fill_white
        
        # Base fields
        ws.cell(row=row_idx, column=1, value=row['question'])
        ws.cell(row=row_idx, column=2, value=row['expected_answer'])
        ws.cell(row=row_idx, column=3, value=row['actual_answer'])
        
        # Overall score with color coding
        overall_score = row['overall_score']
        score_cell = ws.cell(row=row_idx, column=4, value=overall_score)
        score_cell.fill = get_score_fill(overall_score)
        score_cell.font = get_score_font(overall_score)
        score_cell.alignment = Alignment(horizontal="center")
        
        # Parameter scores
        parameter_scores = row['parameter_scores']
        col_offset = 5
        
        for i in range(5):
            base_col = col_offset + (i * 4)
            if i < len(parameter_scores):
                param_score = parameter_scores[i]
                ws.cell(row=row_idx, column=base_col, value=param_score.parameter.name if param_score.parameter else f'Parameter {param_score.parameter_id}')
                
                # Parameter score with color coding
                param_score_val = param_score.score or 0
                param_score_cell = ws.cell(row=row_idx, column=base_col + 1, value=param_score_val)
                param_score_cell.fill = get_score_fill(param_score_val)
                param_score_cell.font = get_score_font(param_score_val)
                param_score_cell.alignment = Alignment(horizontal="center")
                
                ws.cell(row=row_idx, column=base_col + 2, value=param_score.weight_used or 0)
                ws.cell(row=row_idx, column=base_col + 3, value=param_score.reasoning or '')
        
        # Apply row fill and borders to all cells in row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col)
            if col not in [4] + [col_offset + (i * 4) + 1 for i in range(5)]:  # Don't override score colors
                cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    # Set column widths
    column_widths = {
        1: 40,   # Question
        2: 40,   # Expected Answer
        3: 40,   # Actual Answer
        4: 15,   # Overall Score
    }
    # Parameter columns
    for i in range(5):
        base = 5 + (i * 4)
        column_widths[base] = 20      # Name
        column_widths[base + 1] = 12  # Score
        column_widths[base + 2] = 12  # Weight
        column_widths[base + 3] = 35  # Reasoning
    
    for col, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Create summary sheet
    summary_ws = wb.create_sheet(title="Summary")
    
    # Summary styling
    summary_header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    summary_header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Summary data
    total_questions = len(rows_data)
    avg_score = sum(r['overall_score'] for r in rows_data) / total_questions if total_questions > 0 else 0
    high_scores = sum(1 for r in rows_data if r['overall_score'] >= 80)
    medium_scores = sum(1 for r in rows_data if 40 <= r['overall_score'] < 80)
    low_scores = sum(1 for r in rows_data if r['overall_score'] < 40)
    
    # Format date with timezone
    if test_run.created_at:
        try:
            # Convert to user's timezone
            user_tz = ZoneInfo(timezone)
            # Ensure created_at is timezone-aware (assume UTC if naive)
            if test_run.created_at.tzinfo is None:
                utc_time = test_run.created_at.replace(tzinfo=ZoneInfo("UTC"))
            else:
                utc_time = test_run.created_at
            local_time = utc_time.astimezone(user_tz)
            test_date = local_time.strftime("%Y-%m-%d %H:%M:%S %Z")
        except Exception:
            # Fallback to UTC if timezone is invalid
            test_date = test_run.created_at.strftime("%Y-%m-%d %H:%M:%S UTC")
    else:
        test_date = "N/A"
    
    # Determine flow or playbook display
    if test_run.playbook_id:
        flow_or_playbook_label = "Playbook"
        flow_or_playbook_value = test_run.playbook_display_name or test_run.playbook_id or "N/A"
    else:
        flow_or_playbook_label = "Start Flow"
        flow_or_playbook_value = test_run.flow_display_name or test_run.flow_name or "Default Start Flow"
    
    # Format pre/post prompt messages - show actual content
    pre_prompts = test_run.pre_prompt_messages or []
    post_prompts = test_run.post_prompt_messages or []
    
    # Extract message content from the message objects
    def format_messages(messages):
        if not messages:
            return "None"
        message_texts = []
        for i, msg in enumerate(messages, 1):
            if isinstance(msg, dict):
                text = msg.get('text', msg.get('content', str(msg)))
            else:
                text = str(msg)
            message_texts.append(f"{i}. {text}")
        return "\n".join(message_texts)
    
    pre_prompt_display = format_messages(pre_prompts)
    post_prompt_display = format_messages(post_prompts)
    
    summary_data = [
        ("Test Run Summary", ""),
        ("", ""),
        ("Test Run Name", test_run.name),
        ("Date", test_date),
        ("", ""),
        ("Configuration", ""),
        ("Agent", test_run.agent_display_name or test_run.agent_name or "N/A"),
        (flow_or_playbook_label, flow_or_playbook_value),
        ("Evaluation Model", test_run.evaluation_model_id or "N/A"),
        ("Batch Size", test_run.batch_size or 10),
        ("Webhooks", "Enabled" if test_run.enable_webhook else "Disabled"),
        ("Pre-Prompt Messages", pre_prompt_display),
        ("Post-Prompt Messages", post_prompt_display),
        ("", ""),
        ("Results Summary", ""),
        ("Total Questions", total_questions),
        ("Average Score", f"{avg_score:.1f}%"),
        ("", ""),
        ("Score Distribution", ""),
        ("High (≥80%)", f"{high_scores} ({high_scores/total_questions*100:.1f}%)" if total_questions > 0 else "0"),
        ("Medium (40-79%)", f"{medium_scores} ({medium_scores/total_questions*100:.1f}%)" if total_questions > 0 else "0"),
        ("Low (<40%)", f"{low_scores} ({low_scores/total_questions*100:.1f}%)" if total_questions > 0 else "0"),
    ]
    
    for row_idx, (label, value) in enumerate(summary_data, 1):
        label_cell = summary_ws.cell(row=row_idx, column=1, value=label)
        value_cell = summary_ws.cell(row=row_idx, column=2, value=value)
        
        if row_idx == 1:
            label_cell.font = Font(bold=True, size=14, color="1E3A5F")
        elif label in ["Configuration", "Results Summary", "Score Distribution"]:
            label_cell.font = Font(bold=True, size=11)
        elif label == "Average Score":
            value_cell.fill = get_score_fill(avg_score)
            value_cell.font = get_score_font(avg_score)
    
    summary_ws.column_dimensions['A'].width = 25
    summary_ws.column_dimensions['B'].width = 40
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"test-run-{test_run.id}-{safe_name}-results.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# Keep old endpoint for backward compatibility
@router.get("/{test_run_id}/export-csv")
async def export_test_run_csv(
    test_run_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export test run results to CSV format (legacy endpoint - use /export?format=csv instead)."""
    return await export_test_run(test_run_id, format="csv", db=db, current_user=current_user)


@router.get("/models/status")
async def get_models_status(current_user: User = Depends(get_current_user)):
    """Get the current status of the model cache."""
    from app.services.model_cache_service import model_cache_service
    
    try:
        models = model_cache_service.cached_models
        return {
            "status": "success" if models else "empty",
            "count": len(models),
            "last_refresh": model_cache_service.last_refresh.isoformat() if model_cache_service.last_refresh else None,
            "cache_valid": model_cache_service._is_cache_valid()
        }
    except Exception as e:
        return {
            "status": "error",
            "error": str(e),
            "count": 0,
            "last_refresh": None,
            "cache_valid": False
        }


@router.post("/models/refresh")
async def refresh_models(current_user: User = Depends(get_current_user)):
    """Force refresh the model cache."""
    from app.services.model_cache_service import model_cache_service
    
    try:
        models = await model_cache_service.refresh_model_cache()
        return {
            "status": "success",
            "message": f"Successfully refreshed {len(models)} models",
            "count": len(models),
            "models": models
        }
    except Exception as e:
        logger.error(f"Model refresh failed: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to refresh models: {str(e)}"
        )
